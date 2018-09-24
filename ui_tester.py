from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
import os
from openpyxl import load_workbook
from openpyxl.styles import colors
import fnmatch
from openpyxl.styles import Font, Color
import time
from selenium.common.exceptions import NoSuchElementException


class EasyFlow():
    def __init__(self):
        self.browser = webdriver.Ie('IEDriverServer.exe')
        self.wait = WebDriverWait(self.browser, 30)

    """登錄"""

    def login(self, login_url, user_name, user_password, lang_switch):
        # login頁面
        self.browser.get(login_url)
        self.find_element('//select[@id="ddlLanguage"]/option[text()="' +
                          lang_switch + '"]').click()
        self.find_element('//input[@id="txtUserId"]').send_keys(user_name)
        self.find_element('//input[@id="txtPassword"]').send_keys(
            user_password)
        self.find_element('//input[@id="txtPassword"]').send_keys(Keys.ENTER)

    """啟動流程測試模式"""

    def start_simulation_mode(self, user_id):
        # 點擊【啟動流程測試模式】連結
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmNavigator"]')
        script_name = self.find_element(
            '//a[text()="啟動流程測試模式 "]').get_attribute('href')
        self.browser.execute_script(script_name)

        # 點擊【啟動流程測試模式】多選框
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]')
        self.find_element('//input[@id="chkFireValidationMode"]').click()

        # 點擊【選取人員】按鈕
        self.find_element('//input[@id="btnChooseUser"]').click()

        # 儲存視窗定位
        self.store_window_position()

        # 切換角色
        self.browser.switch_to.window(self.sub_window)
        self.find_element(
            '//select[@id="ddlConditionName"]/option[@value="userId"]').click(
            )
        self.find_element('//input[@id="txtConditionValue"]').send_keys(
            user_id)
        self.find_element('//input[@id="btnStartQuerying"]').click()
        self.find_element(
            '(//table[@class="text11"]//tr[@class="ListBodyClass_odd"])[1]'
        ).click()
        self.browser.switch_to.window(self.main_window)

    def invoke_process(self, process_name):
        # 點擊【發起流程】
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmNavigator"]')
        script_name = self.find_element('//a[text()="發起流程 "]').get_attribute(
            'href')
        self.browser.execute_script(script_name)

        # 點擊【某某流程】
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]')
        script_name = self.find_element('//a[text()="' + process_name +
                                        ' "]').get_attribute('href')
        self.browser.execute_script(script_name)
        self.switch_frame('//iframe[@name="ifmAppLocation"]')

    def find_element(self, element_xpath):
        item = self.wait.until(
            EC.element_to_be_clickable((By.XPATH, element_xpath)))
        return item

    def switch_frame(self, element_xpath):
        self.wait.until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH,
                                                        element_xpath)))
        return self

    def upload_attachment(self, element_xpath, file_path, file_description):
        self.find_element(element_xpath).send_keys(Keys.ENTER)

        # 儲存視窗定位
        self.store_window_position()

        self.browser.switch_to.window(self.sub_window)
        self.find_element('//input[@name="fileDocument"]').send_keys(file_path)
        self.find_element('//textarea[@id="txaDocumentDesc"]').send_keys(
            file_description)
        self.find_element('//input[@id="btnUploadDocument"]').click()

        self.browser.switch_to.alert.accept()
        self.browser.close()

        self.browser.switch_to.window(self.main_window)
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]'
                          ).switch_frame('//iframe[@name="ifmAppLocation"]')

    def add_countersign(self, element_xpath, user_id):
        self.find_element(element_xpath).click()

        # 儲存視窗定位
        self.store_window_position()

        # 會簽人員
        self.browser.switch_to.window(self.sub_window)
        self.find_element(
            '//select[@id="ddlConditionName"]/option[@value="userId"]').click(
            )
        self.find_element('//input[@id="txtConditionValue"]').send_keys(
            user_id)
        self.find_element('//input[@id="btnStartQuerying"]').click()
        self.find_element(
            '//table[@class="text11"]//input[@id="chkChosenOptions"]').click()
        self.find_element('//input[@id="btnReturnValues"]').click()

        self.browser.switch_to.window(self.main_window)
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]'
                          ).switch_frame('//iframe[@name="ifmAppLocation"]')

    """儲存父子視窗定位以跳轉操作"""

    def store_window_position(self):
        while True:
            if len(self.browser.window_handles) == 2:
                self.main_window = self.browser.window_handles[0]
                self.sub_window = self.browser.window_handles[1]
                break

    def dispatch_process(self):
        self.click_toolbar_button('派送')
        self.browser.switch_to.alert.accept()

    def change_assignee(self):
        # 儲存視窗定位
        self.store_window_position()
        # 列出工作受託者
        self.browser.switch_to.window(self.sub_window)
        script_name = self.find_element(
            '//input[@name="rdoWorkAssigneeOID"]').get_attribute('onclick')
        self.browser.execute_script(script_name)

        self.browser.switch_to.window(self.main_window)
        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]'
                          ).switch_frame('//iframe[@name="ifmAppLocation"]')

    def get_value_by_name(self, param_values, param_name):
        values = param_values.split(',')
        for value in values:
            if param_name not in value:
                continue
            else:
                return value.split('=')[1]

        return ''

    def run_testcases(self, testcases_path):
        for dir_path, dir_names, file_names in os.walk(testcases_path):
            for file_name in file_names:
                if '(已通過)' not in file_name and fnmatch.fnmatch(
                        file_name, '*.xlsx'):
                    file_path = os.path.join(dir_path, file_name)
                    workbook = load_workbook(file_path)
                    test_passed = True
                    for sheet in workbook._sheets:
                        for index, row in enumerate(sheet.iter_rows()):
                            current_row = index + 1
                            action_name = str(
                                sheet.cell(row=current_row, column=1).value)
                            target_condition = str(
                                sheet.cell(row=current_row, column=2).value)
                            setting_value = str(
                                sheet.cell(row=current_row, column=3).value)
                            if (action_name == '登入'):
                                account = self.get_value_by_name(
                                    setting_value, '帳號')
                                password = self.get_value_by_name(
                                    setting_value, '密碼')
                                lang_switch = self.get_value_by_name(
                                    setting_value, '語系')
                                self.login(target_condition, account, password,
                                           lang_switch)
                            if (action_name == '模擬'):
                                self.start_simulation_mode(target_condition)
                            if (action_name == '發起流程'):
                                self.invoke_process(target_condition)
                            if (action_name == '點擊'):
                                element_type = self.get_value_by_name(
                                    target_condition, '控件類型')
                                element_name = self.get_value_by_name(
                                    target_condition, '控件名稱')
                                element_xpath = '//' + element_type + '[@name="' + element_name + '"'
                                element_text = self.get_value_by_name(
                                    setting_value, '文字')
                                element_value = self.get_value_by_name(
                                    setting_value, '值')
                                if (element_text != ''):
                                    element_xpath += ' and text()="' + element_text + '"'
                                if (element_value != ''):
                                    element_xpath += ' and @value="' + element_value + '"'
                                element_xpath += ']'
                                self.find_element(element_xpath).click()
                            if (action_name == '輸入'):
                                element_type = self.get_value_by_name(
                                    target_condition, '控件類型')
                                element_name = self.get_value_by_name(
                                    target_condition, '控件名稱')
                                element_xpath = '//' + element_type + '[@name="' + element_name + '"]'
                                element_text = self.get_value_by_name(
                                    setting_value, '文字')
                                element_value = self.get_value_by_name(
                                    setting_value, '值')
                                if (element_text != ''):
                                    self.find_element(element_xpath).send_keys(
                                        element_text)
                                if (element_value != ''):
                                    self.find_element(
                                        element_xpath).value = element_value
                            if (action_name == '附件'):
                                element_name = self.get_value_by_name(
                                    target_condition, '控件名稱')
                                attacment_path = self.get_value_by_name(
                                    setting_value, '路徑')
                                attacment_name = self.get_value_by_name(
                                    setting_value, '名稱')
                                element_xpath = '//input[@name="' + element_name + '"]'
                                self.upload_attachment(element_xpath,
                                                       attacment_path,
                                                       attacment_name)
                            if (action_name == '會簽'):
                                self.add_countersign(target_condition)
                            if (action_name == '派送'):
                                self.dispatch_process()
                            if (action_name == '預期'):
                                if (target_condition == '警告'):
                                    expected = self.get_value_by_name(
                                        setting_value, '文字')
                                    actual = self.browser.switch_to_alert(
                                    ).text
                                    self.browser.switch_to_alert().accept()
                                    next_a_cell = sheet.cell(
                                        row=current_row + 1, column=1)
                                    next_b_cell = sheet.cell(
                                        row=current_row + 1, column=2)
                                    next_c_cell = sheet.cell(
                                        row=current_row + 1, column=3)
                                    afternext_a_cell = sheet.cell(
                                        row=current_row + 2, column=1)
                                    afternext_b_cell = sheet.cell(
                                        row=current_row + 2, column=2)
                                    next_a_cell.value = '實際'
                                    next_b_cell.value = target_condition
                                    next_c_cell.value = '文字=' + actual
                                    afternext_a_cell.value = '測試結果'
                                    if (expected in actual):
                                        afternext_b_cell.value = '符合'
                                        afternext_b_cell.font = Font(
                                            color=colors.GREEN, bold=True)
                                        sheet.sheet_properties.tabColor = '00FF00'
                                    else:
                                        afternext_b_cell.value = '不符合'
                                        afternext_b_cell.font = Font(
                                            color=colors.RED, bold=True)
                                        sheet.sheet_properties.tabColor = 'FF0000'
                                        test_passed = False                                        
                                if (action_name == ''):
                                    break
                    workbook.save(file_path)
                    if (test_passed == True):
                        os.rename(file_path,
                                  os.path.join(dir_path, '(已通過)' + file_name))
        self.browser.quit()

    def open_first_workitem(self):
        script_name = self.find_element(
            '(//tr[@class="ListBodyClass_odd"])[1]//samp').get_attribute(
                'onclick')
        self.browser.execute_script(script_name)

    def accept_workitem(self):
        self.click_toolbar_button('接收')
        self.browser.switch_to.alert.accept()

    def click_toolbar_button(self, action_name):
        actions = {'派送': 'btnInvokeProcess', '接收': 'btnAcceptWorkItem'}

        self.browser.switch_to.default_content()
        self.switch_frame('//iframe[@name="ifmFucntionLocation"]')
        script_name = self.find_element('//img[@id="' +
                                        actions.get(action_name, None) +
                                        '"]').get_attribute('onclick')
        self.browser.execute_script(script_name)

    def save_snapshot(self, file_path):
        self.browser.get_screenshot_as_file(file_path)

    def set_datetimepicker_value(self, element_name, datetime_string):
        datetime_control = self.find_element('//input[@name="' + element_name +
                                             '"]')
        readonly_attribute = datetime_control.get_attribute('readonly')
        disabled_attribute = datetime_control.get_attribute('disabled')
        if (readonly_attribute == 'readonly'):
            self.browser.execute_script('document.getElementsByName("' +
                                        element_name +
                                        '")[0].removeAttribute("readonly");')
        if (disabled_attribute == 'disabled'):
            self.browser.execute_script('document.getElementsByName("' +
                                        element_name +
                                        '")[0].removeAttribute("disabled");')
        datetime_control.clear()
        datetime_control.send_keys(datetime_string)
        if (readonly_attribute == 'readonly'):
            self.browser.execute_script('document.getElementsByName("' +
                                        element_name +
                                        '")[0].addAttribute("readonly");')
        if (disabled_attribute == 'disabled'):
            self.browser.execute_script('document.getElementsByName("' +
                                        element_name +
                                        '")[0].addAttribute("disabled");')
